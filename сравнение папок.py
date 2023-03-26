import os
import binascii
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def get_crc32(file_path):
    with open(file_path, 'rb') as f:
        buf = f.read()
        crc32 = format(binascii.crc32(buf), '08X')
    return crc32

def get_files_data(folder):
    result = {}
    for root, _, files in os.walk(folder):
        for file in files:
            file_path = os.path.join(root, file)
            relative_path = os.path.relpath(file_path, folder)
            size = os.path.getsize(file_path)
            crc32 = get_crc32(file_path)
            result[relative_path] = (size, crc32)
    return result

def save_xlsx(file_data1, file_data2, output_path):
    wb = Workbook()
    ws = wb.active

    ws.append(["Path 1", "Size 1", "CRC-32 1", "Path 2", "Size 2", "CRC-32 2"])

    red_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for key in sorted(set(file_data1.keys()) | set(file_data2.keys())):
        data1 = file_data1.get(key, ('-', '-'))
        data2 = file_data2.get(key, ('-', '-'))

        row = [key, data1[0], data1[1], key, data2[0], data2[1]]
        ws.append(row)

        if data1 != data2:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    wb.save(output_path)

if __name__ == '__main__':
    folder1 = input('Введите путь к первой папке: ').strip()
    folder2 = input('Введите путь ко второй папке: ').strip()

    file_data1 = get_files_data(folder1)
    file_data2 = get_files_data(folder2)

    output_path = 'output.xlsx'
    save_xlsx(file_data1, file_data2, output_path)

    print(f'Файл {output_path} успешно сохранен.')
