import os
import binascii
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def get_crc32(file_path):
    with open(file_path, 'rb') as f:
        buf = f.read()
        crc32 = format(binascii.crc32(buf), '08X')
    return crc32


def get_files_data(folder):
    result = {}
    for root, _, files in os.walk(folder):
        for file in files:
            file_path = os.path.join(root, file)
            relative_path = os.path.relpath(file_path, folder)
            size = os.path.getsize(file_path)
            crc32 = get_crc32(file_path)
            result[relative_path] = (size, crc32)
    return result


def save_xlsx(file_data1, file_data2, output_path):
    wb = Workbook()
    ws = wb.active

    ws.append(["Path 1", "Size 1", "CRC-32 1", "Path 2", "Size 2", "CRC-32 2"])

    red_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for key in sorted(set(file_data1.keys()) | set(file_data2.keys())):
        data1 = file_data1.get(key, ('-', '-'))
        data2 = file_data2.get(key, ('-', '-'))

        row = [key, data1[0], data1[1], key, data2[0], data2[1]]
        ws.append(row)

        if data1 != data2:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    wb.save(output_path)


def run_comparison(folder1, folder2):
    if folder1 and folder2:
        file_data1 = get_files_data(folder1)
        file_data2 = get_files_data(folder2)

        output_path = 'output.xlsx'
        save_xlsx(file_data1, file_data2, output_path)

        print(f'Файл {output_path} успешно сохранен.')
    else:
        print("Вы не выбрали папки. Завершение работы.")


def select_folder(entry):
    folder = filedialog.askdirectory()
    if folder:
        entry.delete(0, tk.END)
        entry.insert(0, folder)


def main():
    root = tk.Tk()
    root.title("Сравнение папок")

    frame = ttk.Frame(root, padding="10")
    frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    folder1_label = ttk.Label(frame, text="Папка 1:")
    folder1_label.grid(row=0, column=0, sticky=tk.W)

    folder1_entry = ttk.Entry(frame, width=50)
    folder1_entry.grid(row=0, column=1, sticky=(tk.W, tk.E))

    folder1_button = ttk.Button(frame, text="Выбрать", command=lambda: select_folder(folder1_entry))
    folder1_button.grid(row=0, column=2, sticky=tk.W)

    folder2_label = ttk.Label(frame, text="Папка 2:")
    folder2_label.grid(row=1, column=0, sticky=tk.W)

    folder2_entry = ttk.Entry(frame, width=50)
    folder2_entry.grid(row=1, column=1, sticky=(tk.W, tk.E))

    folder2_button = ttk.Button(frame, text="Выбрать", command=lambda: select_folder(folder2_entry))
    folder2_button.grid(row=1, column=2, sticky=tk.W)

    compare_button = ttk.Button(frame, text="Сравнить", command=lambda: run_comparison(folder1_entry.get(), folder2_entry.get()))
    compare_button.grid(row=2, column=0, columnspan=3, pady=(10, 0))

    root.mainloop()


if __name__ == '__main__':
    main()

   